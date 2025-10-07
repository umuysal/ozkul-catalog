from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict, Any
import os, shutil

def import_from_excel(xlsx_path: str, upload_dir: str):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active  # assume first sheet
    # Identify header row
    headers = {}
    for j, cell in enumerate(ws[1], start=1):
        headers[cell.value.strip().lower() if isinstance(cell.value, str) else f"col{j}"] = j

    def get(col_name, row):
        j = headers.get(col_name)
        return ws.cell(row=row, column=j).value if j else None

    # Map images to row via anchor row
    image_map = {}
    for img in ws._images:
        try:
            row = img.anchor._from.row + 1  # openpyxl anchor is 0-indexed
        except Exception:
            continue
        image_map.setdefault(row, []).append(img)

    rows: List[Dict[str, Any]] = []
    os.makedirs(upload_dir, exist_ok=True)

    for i in range(2, ws.max_row+1):
        sku = get('sku', i) or get('stok_kodu', i) or f"SKU-{i-1:05d}"
        name = get('name', i) or get('ürün adı', i) or get('urun_adi', i) or f"Ürün {i-1}"
        description = get('description', i) or get('açıklama', i) or ''
        price = get('price', i) or get('fiyat', i)

        paths = []
        for idx, img in enumerate(image_map.get(i, [])):
            ext = '.png'
            fname = f"{sku}_{idx}{ext}"
            out_path = os.path.join(upload_dir, fname)
            try:
                img_ref = img._data()
                with open(out_path, 'wb') as f:
                    f.write(img_ref())
                paths.append(out_path)
            except Exception:
                # fallback: skip
                pass

        rows.append({
            'sku': str(sku),
            'name': str(name),
            'description': str(description) if description else '',
            'price': float(price) if price is not None else None,
            'images': paths
        })
    return rows
